from ast import If, main
from collections import namedtuple
import altair as alt
import math
import pandas as pd
import streamlit as st
import difflib
from PIL import Image
import imagehash
import cv2
import matplotlib.pyplot as plt
import glob
import os
from io import BytesIO
import requests
import openpyxl
from openpyxl import Workbook
import numpy as np
import pandas as pd
import xlrd
import xlsxwriter
from datetime import datetime


"""
# Welcome to Streamlit!

Edit `/streamlit_app.py` to customize this app to your heart's desire :heart:

If you have any questions, checkout our [documentation](https://docs.streamlit.io) and [community
forums](https://discuss.streamlit.io).

In the meantime, below is an example of what you can do with just a few lines of code:
"""


list_score = []
list_mask = []
list_all = []

def smart_crop(path):
    img = cv2.imread(path)
    gry = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gry,(3,3), 0)
    th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY)[1]
    coords = cv2.findNonZero(th)
    x,y,w,h = cv2.boundingRect(coords)
    image = cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 0)
    crop_img = image[y:y+h, x:x+w]
    crop = Image.fromarray(crop_img)
    crop = crop.convert("RGB")
    w, h = crop.size
    print(str(path))
    result_resize = float(h/w)
    if result_resize < 1.675:
        up_points = (400,400)
        crop_imag = crop.resize(up_points)
    else:
        up_points = (200,400)
        crop_imag = crop.resize(up_points)
    return crop_imag

def path_cut(path_to_file):
  path_name = os.path.normpath(path_to_file)
  text_path = str(path_name.split(os.sep)[3] + "_" + path_name.split(os.sep)[4])
  return text_path


def get_img(image, size=(100, 100)):
    img = Image.open(image)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format="png")
    temp.seek(0)
    return Image.open(temp)


def insert_row(ws, image_1, image_2, name, num, num_2,num_3,num_4,num_5, size=(200,200)):
    img_1 = openpyxl.drawing.image.Image(get_img(image_1, size=size))
    img_2 = openpyxl.drawing.image.Image(get_img(image_2, size=size))
    row_num = ws.max_row + 1
    cell_addr_1 = f"A{row_num}"
    img_1.anchor = cell_addr_1
    ws.add_image(img_1)
    
    row_num = ws.max_row + 1
    cell_addr_2 = f"B{row_num}"
    img_2.anchor = cell_addr_2
    ws.add_image(img_2)
    ws[f"C{row_num}"] = name
    ws[f"D{row_num}"] = num
    ws[f"E{row_num}"] = num_2
    ws[f"F{row_num}"] = num_3
    ws[f"G{row_num}"] = num_4
    ws[f"H{row_num}"] = num_5
    ws.row_dimensions[row_num].height = int(size[1] * .8)
    ws.column_dimensions["A"].width = int(size[0] * .2)
    ws.column_dimensions["B"].width = int(size[0] * .2)


def hash_check_4(path_1,path_2,hash_size):
    img1 = smart_crop(path_1)
    img2 = smart_crop(path_2)

    # hash_a = imagehash.average_hash(img1,hash_size)
    # otherhash = imagehash.average_hash(img2,hash_size)
    # delta_1 = hash_a - otherhash

    # hash_p = imagehash.phash(img1,hash_size)
    # otherhash_p = imagehash.phash(img2,hash_size)
    # delta_2 = hash_p - otherhash_p

    hash_d = imagehash.dhash(img1,hash_size)
    otherhash_d = imagehash.dhash(img2,hash_size)
    delta_4 = hash_d -  otherhash_d

    hash_dhash_vertical = imagehash.dhash_vertical(img1,hash_size)
    otherhash_dhash_vertical = imagehash.dhash_vertical(img2,hash_size)
    delta_5 = hash_dhash_vertical -  otherhash_dhash_vertical

    av = ( delta_5 + delta_4)/2 
    print("AVG: ",av)
    plt.subplot(121)
    plt.imshow(img1)
    plt.subplot(122)
    plt.imshow(img2)

    plt.show()
    return av, delta_5, delta_4

def check_masks_l(path_one_mask, path_all_mask,hash_count):
    size = (200, 300)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Img_1"
    ws["B1"] = "Img_2"
    ws["C1"] = "Filename"
    ws["D1"] = "AVG_HASH" + str(hash_count)+ 'x' + str(hash_count) + " L"
    ws["E1"] = "d_hash_v"
    ws["F1"] = "d_hash"
    ws["G1"] = "p_hash"
    ws["H1"] = "a_hash"
    for folder_all_mask in glob.glob(path_all_mask):
        
        score , delta_5, delta_4 = hash_check(path_one_mask, folder_all_mask,hash_count)
        list_score.append(score)
        list_mask.append(folder_all_mask)
        insert_row(ws,  path_one_mask, folder_all_mask, os.path.basename(folder_all_mask),score,delta_5, delta_4, delta_3, delta_2, size=size)
    wb.save('/content/result/new_test_L' + str(hash_count)+ 'x' + str(hash_count) +'.xlsx')
    wb.close()

def load_image(image_file):
	img = Image.open(image_file)
	return img


if __name__ == '__main__':
    st.title("File Upload Tutorial")

	choice = st.sidebar.selectbox("Menu")

	if choice == "Image":
		st.subheader("Image")
    	image_file = st.file_uploader("Upload Images", type=["png","jpg","jpeg"])

		if image_file is not None:
      		file_details = {"filename":image_file.name,     
                      "filetype":image_file.type,
                      "filesize":image_file.size}
      		st.write(file_details)
      		st.image(load_image(image_file),width=250)
